from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class Formatter:

    @staticmethod
    def format_excel(file_path):

        wb = load_workbook(file_path)

        for sheet in wb.sheetnames:

            ws = wb[sheet]

            for cell in ws[1]:

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    start_color="4F81BD",
                    end_color="4F81BD",
                    fill_type="solid"
                )

        wb.save(file_path)

        print("Formatting applied")